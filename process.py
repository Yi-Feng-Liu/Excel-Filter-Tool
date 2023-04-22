import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment, numbers
import copy
import time


class Judge_Metabolic_Syndrome:
    def __init__(self, io, tab, select_years, save_sheet_name, years_text=None, from_summary=False):
        self.io = io
        self.tab = tab
        self.dst_worksheet = save_sheet_name
        self.select_years = select_years
        self.years_text = years_text
        self.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        self.font = Font(name='Calibri', color='FF0000')
        self.font_type = Font(name='Calibri')
        self.from_summary = from_summary
        self.gender_dict = {'gender': 5}
        self.column_dict={
            'waistline':9,
            'systolic':10,
            'diastolic':11,
            'glucose':12,
            'triglycerides':14,
            'hdlc':15
        }
        self.standard_dict={
            'waistline': 90,
            'systolic': 130,
            'diastolic': 85,
            'glucose':100,
            'triglycerides':150,
            'hdlc': 40
        }   


    def change_date_time(self, worksheet, number_of_column):
        """Remove hours:minute:second format of the datetime 

        Args:
            worksheet : excel worksheet
            number_of_column : If column name is G, the number of column is 7, etc.

        Returns:
            _worksheet: 
        """
        # 變更時間格式
        for row in worksheet.iter_rows(min_row=2, min_col=number_of_column, max_col=number_of_column):
            for cell in row:
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
        return worksheet


    def place_center(self, worksheet):
        align = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = align
        return worksheet


    def set_specific_column_format(self, worksheet:str, eng_column:str, width=15, only_change_width=False):
        """set the specific column format.

        Args:
            worksheet (str): Excel worksheet
            eng_column (str): like 'A' or 'G' column
        """
        if only_change_width==True:
            worksheet.column_dimensions[eng_column].width = width
        else:
            worksheet.column_dimensions[eng_column].width = width
            worksheet[f'{eng_column}1'].fill = self.fill
            worksheet[f'{eng_column}1'].font = self.font
        return worksheet


    def change_font_color_format(self, cell):
        """Change font color 

        Args:
            cell : the cell coordinate
        """
        cell.font = self.font
    
    def change_font_type_format(self, worksheet):
        """Change font type 

        Args:
            cell : the cell coordinate
        """
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                # 設置儲存格的字體
                cell.font = self.font_type
        return worksheet

    def label_over_standard(self, worksheet):
        """Use to label the cell, if cell's value exceed the standard

        Args:
            worksheet: the excel work sheet 

        Returns:
            worksheet
        """
        worksheet = self.change_font_type_format(worksheet)

        for row in worksheet.iter_rows(min_row=2):
            people_name = row[1]
            gender = row[self.gender_dict['gender']] 
            over_standard_cnt = 0
            if gender.value == '男':
                for key, value in self.column_dict.items():
                    if row[value].value is None:
                        continue
                    if row[value].value == '無資料':
                        continue
                    if isinstance(row[value].value, str):
                        row[value].value = float(row[value].value)
                        if key == 'hdlc' and row[value].value < self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key != 'hdlc' and row[value].value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)
                    else:
                        if key == 'waistline' and row[value].value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key =='hdlc' and row[value].value < self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key != 'hdlc' and row[value].value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)

            elif gender.value == '女':
                for key, value in self.column_dict.items():
                    if row[value].value is None:
                        continue
                    if row[value].value == '無資料':
                        continue
                    if isinstance(row[value].value, str):
                        row[value].value = float(row[value].value.split('(')[0])
                        if key == 'waistline' and row[value].value >= self.standard_dict[key]-10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key =='hdlc' and row[value].value < self.standard_dict[key]+10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key != 'hdlc' and row[value].value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)
                    else:
                        if key == 'waistline' and row[value].value >= self.standard_dict[key]-10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key =='hdlc' and row[value].value < self.standard_dict[key]+10:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        elif key != 'hdlc' and row[value].value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                            self.change_font_color_format(row[value])
                        if over_standard_cnt >= 3:
                            self.change_font_color_format(people_name)
        return worksheet
    

    def copy_title_format(self, ws1, ws2):
        for row in ws1.iter_rows(min_row=1, max_row=1):
            for cell in row:
                # copy cell
                ws2[cell.coordinate].font = copy.copy(cell.font)
                ws2[cell.coordinate].fill = copy.copy(cell.fill)
                ws2[cell.coordinate].border = copy.copy(cell.border)
                ws2[cell.coordinate].alignment = copy.copy(cell.alignment)
                ws2[cell.coordinate].number_format = copy.copy(cell.number_format)
                ws2.row_dimensions[cell.row].height = copy.copy(ws1.row_dimensions[cell.row].height)
                ws2.column_dimensions[cell.column_letter].width = copy.copy(ws1.column_dimensions[cell.column_letter].width)
        return ws2
    
    def copy_format_from_sheet1(self):
        """Copy the original sheet header format to specific sheet

        Including cell's fill, font, color, alignment, dimensions.
        """
        workbook = openpyxl.load_workbook(self.io)

        ws1 = workbook[self.tab]
        ws2 = workbook[self.dst_worksheet]
        # copy format sheet1 header to sheet2 header
        ws2 = self.copy_title_format(ws1=ws1, ws2=ws2)
        ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='U', width=13, only_change_width=False)
        if self.from_summary==True:
            start_column = 71 # G
            end_column = 81 #Q
            for i in range(start_column, end_column+1):
                ws2 = self.set_specific_column_format(worksheet=ws2, eng_column=chr(i), width=15, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='E', width=42, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='R', width=45, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='S', width=35, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='T', width=24, only_change_width=True)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='V', width=13)
            ws2 = self.set_specific_column_format(worksheet=ws2, eng_column='W', width=13)
        ws2 = self.change_date_time(worksheet=ws2, number_of_column=7)
        ws2 = self.place_center(worksheet=ws2)

        # label_over_standard worksheet
        # ws1 = self.label_over_standard(worksheet=ws1)
        # only process new sheet
        ws2 = self.label_over_standard(worksheet=ws2)
        workbook.save(self.io)
        print("saved")


    def read_file(self):
        if self.from_summary==False:
            df = pd.read_excel(self.io, sheet_name='健檢資料', engine='openpyxl')
        else:
            df = pd.read_excel(self.save_file_path, sheet_name='健檢資料', engine='openpyxl')
        return df
    
    
    def process_Metabolic_Syndrome(self, df:pd.DataFrame):
        """Filter file and add new column named 超過標準數,

        Args:
            df (pd.DataFrame): Original dataframe
        Returns:
            dataframe
        """
        df['超過標準數'] = 0
        if self.from_summary==False:
            df = df[df['年度代碼'].str.startswith(self.select_years)]
        else:
            df = df[df['年度代碼'].str.startswith(self.years_text)]
             
        for i in range(len(df.index)):
            name = df.iloc[i, 1] 
            gender = df.iloc[i, self.gender_dict['gender']] 
            over_standard_cnt = 0
            if gender == '男':
                for key, value in self.column_dict.items():
                    df_value = df.iloc[i, value]
                    if pd.isna(df_value):
                        continue
                    if df_value=='無資料':
                        df_value = ''
                        continue
                    if isinstance(df_value, str):
                        df_value = float(df_value)
                        if key == 'hdlc' and df_value < self.standard_dict[key]:
                            over_standard_cnt += 1
                        elif key != 'hdlc' and df_value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                        df.iloc[i, len(df.columns)-1] = over_standard_cnt
                    else:
                        if key == 'hdlc' and df_value < self.standard_dict[key]:
                            over_standard_cnt += 1
                        elif key != 'hdlc' and df_value >= self.standard_dict[key]:
                            over_standard_cnt += 1
                        df.iloc[i, len(df.columns)-1] = over_standard_cnt
            elif gender == '女':
                for key, value in self.column_dict.items():
                    df_value = df.iloc[i, value]
                    if pd.isna(df_value):
                        continue
                    if df_value=='無資料':
                        df_value = ''
                        continue
                    if isinstance(df_value, str):
                        df_value = float(df_value.split('(')[0])
                        if key == 'waistline' and df_value >= self.standard_dict[key]-10:
                            over_standard_cnt += 1
                        elif key =='hdlc' and df_value < self.standard_dict[key]+10:
                            over_standard_cnt += 1
                        elif key != 'hdlc' and df_value >= self.standard_dict[key]:
                            over_standard_cnt += 1 
                        df.iloc[i, len(df.columns)-1] = over_standard_cnt
                    else:
                        if key == 'waistline' and df_value >= self.standard_dict[key]-10:
                            over_standard_cnt += 1
                        elif key =='hdlc' and df_value < self.standard_dict[key]+10:
                            over_standard_cnt += 1
                        elif key != 'hdlc' and df_value >= self.standard_dict[key]:
                            over_standard_cnt += 1 
                        df.iloc[i, len(df.columns)-1] = over_standard_cnt
        df = df.sort_values(by=['超過標準數'], ascending=False)
        return df
        
    
    def save_file_and_copy_title(self, df):
        # 建立一個新的 ExcelWriter 物件
        writer = pd.ExcelWriter(self.io, mode='a', engine='openpyxl', if_sheet_exists='replace')
        df.to_excel(writer, sheet_name=self.dst_worksheet, index=False)
        writer.close() 


    def main_procesdure(self):
        df = self.read_file()
        df = self.process_Metabolic_Syndrome(df)
        self.save_file_and_copy_title(df)
        self.copy_format_from_sheet1()


class Metabolic_Syndrome_From_Summary(Judge_Metabolic_Syndrome):
    def __init__(self, io, tab, save_sheet_name, years_text, from_summary=True):
        super().__init__(io, tab, save_sheet_name, years_text)
        self.io = io
        self.years_text = years_text
        self.save_sheet_name = save_sheet_name
        self.from_summary = from_summary

    def change_column_name(self):
        self.df = pd.read_excel(self.io, engine='openpyxl')
        self.df['年度代碼'] = self.years_text
        self.df['部門代號'] = 'X001'
        self.df['健檢過程備註說明'] = ''
        self.df = self.df.drop(labels=0, axis=0)

        # self.df = self.df.replace('無資料', ' ', inplace=True)

        self.goal_column_name = ['年度代碼', '姓名', '員工編號', '部門代號', '部門名稱', '性別', '出生年月日', '身高_cm', '體重_kg', '腰圍_cm', '收縮壓', '舒張壓', '飯前血醣', '總膽固醇', '三酸甘油脂', '高密度膽固醇', '低密度膽固醇', '抽菸習慣', '既往病歷', '健檢過程備註說明', 'SGOT', 'SGOT']

        specific_column = ['年度代碼', '姓名', '工／學號', '部門代號', '部門／科系', '性別', '生日', '身高', '體重', '腰圍', '收縮壓', '舒張壓', 'AC飯前血糖', 'T-CHO總膽固醇', 'TG三酸甘油脂', 'HDL高密度脂蛋白', 'LDL低密度脂蛋白', '吸菸習慣', '既往病史', '健檢過程備註說明', 'SGOT血清麩酸草酸轉氨脢', 'SGPT血清麩酸丙銅轉氨脢']

        speific_df = self.df.copy()
        speific_df = speific_df[specific_column]
        speific_df_columns = speific_df.columns.to_list()
        for i in range(len(speific_df_columns)):
            speific_df.rename(columns={speific_df_columns[i]: self.goal_column_name[i]}, inplace=True)  
        return speific_df
    
    def main_procesdure(self):
        speific_df = self.change_column_name()
        speific_df = self.process_Metabolic_Syndrome(speific_df)
        self.save_file_and_copy_title(speific_df)
        self.copy_format_from_sheet1()
        print('OK')
        
        
        
    
def main():
    # Metabolic_Syndrome_From_Summary('111 (1).xlsx', '111 年度健檢', 'test_data', years_text='111年度健檢').main_procesdure()
    Judge_Metabolic_Syndrome('test.xlsx', '健檢資料', '111', '123').main_procesdure()

if __name__ == '__main__':
    main()
